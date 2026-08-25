from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from gestion.models import Articulo


class ExcelService:
    """Servicio para exportar artículos a Excel"""
    
    @staticmethod
    def generar_excel():
        """
        Genera un archivo Excel con todos los artículos.
        
        Retorna: BytesIO con el contenido del archivo Excel
        """
        
        # Obtener artículos con select_related para evitar N+1
        articulos = Articulo.objects.select_related(
            'sru_codi',
            'sru_codi__rub_codi'
        ).all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Artículos"
        
        # Definir encabezados
        encabezados = [
            "Código",
            "Nombre",
            "Rubro",
            "SubRubro",
            "Precio Final",
            "IVA (%)"
        ]
        
        # Agregar encabezados a la primera fila
        for col_num, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = encabezado
            # Estilo: negrita, fondo gris, centrado
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Agregar datos
        for row_num, articulo in enumerate(articulos, 2):
            fila_datos = [
                articulo.art_codi,
                articulo.art_nomb,
                articulo.sru_codi.rub_codi.rub_nomb if articulo.sru_codi and articulo.sru_codi.rub_codi else "",
                articulo.sru_codi.sru_nomb if articulo.sru_codi else "",
                float(articulo.art_pfin) if articulo.art_pfin else 0,
                float(articulo.art_tiva) if articulo.art_tiva else 0,
            ]

            for col_num, valor in enumerate(fila_datos, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = valor
                # Alineación y formato
                if col_num in [5, 6]:  # Columnas numéricas (Precio Final, IVA)
                    cell.alignment = Alignment(horizontal="right")
                    if col_num == 5:  # Precio - formato moneda
                        cell.number_format = '$#,##0.00'
                    else:  # IVA - formato decimal
                        cell.number_format = '0.00'
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
        
        # Ajustar ancho de columnas automáticamente
        for col_num in range(1, len(encabezados) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar encabezados
        ws.freeze_panes = "A2"
        
        # Generar BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
